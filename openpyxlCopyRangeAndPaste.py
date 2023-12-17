import openpyxl

# 如果是單純的 Excel 表格運行良好，但是如果有樞紐分析表在工作表裏好像不行
# 將來源(複製)目標範圍存到巢狀列表(list)
# 提供： 起始座標、終止座標、工作表物件
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # 所選列的迴圈
    for i in range(startRow,endRow + 1,1):
        # 每一列加到 list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        # 把結果加到 rangeSelected (巢狀)
        rangeSelected.append(rowSelected)

    return rangeSelected

# 貼到範圍
# 把 copyRange 內容(copiedData)貼到目標 sheet(sheetReceiving)
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


# 開啟 Excel 文件
workbook = openpyxl.load_workbook(r'input.xlsx')

# 选择要复制的工作表
sheet = workbook['工作表1']

# 指定複製和貼上的範圍，全部都轉成數字，從1起算
selectedRange = copyRange(27,3,38,10,sheet)
pastingRange = pasteRange(1,3,12,10,sheet,selectedRange)
# 儲存檔案
workbook.save(r'input.xlsx')

